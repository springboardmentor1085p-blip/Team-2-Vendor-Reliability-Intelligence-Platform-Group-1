import csv
import io
from typing import Literal

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

import models
from database import SessionLocal

# ── reportlab imports ─────────────────────────────────────────────────────────
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

# ── openpyxl imports ──────────────────────────────────────────────────────────
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


router = APIRouter(
    prefix="/reports",
    tags=["Reports"],
)


# ── Dependency ────────────────────────────────────────────────────────────────

def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()


# ── Constants ─────────────────────────────────────────────────────────────────

_COLUMNS = ("po_number", "vendor_id", "status", "order_date", "total_amount")
_HEADERS = ("PO Number", "Vendor ID", "Status", "Order Date", "Total Amount")


# ── Helpers ───────────────────────────────────────────────────────────────────

def _fetch_rows(db: Session) -> list[tuple]:
    """Query purchase_orders and return the five report columns as plain tuples."""
    rows = (
        db.query(
            models.PurchaseOrder.po_number,
            models.PurchaseOrder.vendor_id,
            models.PurchaseOrder.status,
            models.PurchaseOrder.order_date,
            models.PurchaseOrder.total_amount,
        )
        .order_by(models.PurchaseOrder.order_date.desc())
        .all()
    )
    return rows


def _format_row(row) -> tuple:
    """Normalise each cell to a plain string suitable for CSV or PDF."""
    po_number    = row.po_number or ""
    vendor_id    = str(row.vendor_id) if row.vendor_id is not None else ""
    po_status    = row.status.value if hasattr(row.status, "value") else str(row.status)
    order_date   = row.order_date.strftime("%Y-%m-%d") if row.order_date else ""
    total_amount = f"{float(row.total_amount):.2f}" if row.total_amount is not None else ""
    return (po_number, vendor_id, po_status, order_date, total_amount)


# ── CSV builder ───────────────────────────────────────────────────────────────

def _build_csv(rows: list[tuple]) -> io.StringIO:
    buffer = io.StringIO()
    writer = csv.writer(buffer, lineterminator="\n")
    writer.writerow(_HEADERS)
    for row in rows:
        writer.writerow(_format_row(row))
    buffer.seek(0)
    return buffer


# ── PDF builder ───────────────────────────────────────────────────────────────

def _build_pdf(rows: list[tuple]) -> io.BytesIO:
    buffer = io.BytesIO()

    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=1.5 * cm,
        rightMargin=1.5 * cm,
        topMargin=1.5 * cm,
        bottomMargin=1.5 * cm,
        title="Purchase Orders Report",
    )

    styles = getSampleStyleSheet()

    # Title paragraph
    title = Paragraph("Purchase Orders Report", styles["Title"])

    # Table data: header row + data rows
    table_data = [list(_HEADERS)]
    for row in rows:
        table_data.append(list(_format_row(row)))

    col_widths = [5 * cm, 3 * cm, 3.5 * cm, 4 * cm, 4.5 * cm]

    tbl = Table(table_data, colWidths=col_widths, repeatRows=1)
    tbl.setStyle(
        TableStyle([
            # Header row styling
            ("BACKGROUND",   (0, 0), (-1, 0),  colors.HexColor("#1565C0")),
            ("TEXTCOLOR",    (0, 0), (-1, 0),  colors.white),
            ("FONTNAME",     (0, 0), (-1, 0),  "Helvetica-Bold"),
            ("FONTSIZE",     (0, 0), (-1, 0),  10),
            ("ALIGN",        (0, 0), (-1, 0),  "CENTER"),
            ("BOTTOMPADDING",(0, 0), (-1, 0),  8),
            ("TOPPADDING",   (0, 0), (-1, 0),  8),
            # Data rows — alternating row shading
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#E3F2FD")]),
            ("FONTNAME",     (0, 1), (-1, -1),  "Helvetica"),
            ("FONTSIZE",     (0, 1), (-1, -1),  9),
            ("ALIGN",        (1, 1), (-1, -1),  "CENTER"),   # centre all except PO Number
            ("ALIGN",        (0, 1), (0, -1),   "LEFT"),
            ("TOPPADDING",   (0, 1), (-1, -1),  5),
            ("BOTTOMPADDING",(0, 1), (-1, -1),  5),
            # Grid
            ("GRID",         (0, 0), (-1, -1),  0.5, colors.HexColor("#BDBDBD")),
            ("BOX",          (0, 0), (-1, -1),  1,   colors.HexColor("#1565C0")),
        ])
    )

    doc.build([title, Spacer(1, 0.5 * cm), tbl])
    buffer.seek(0)
    return buffer


# ── Excel builder ─────────────────────────────────────────────────────────────

def _build_excel(rows: list[tuple]) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Purchase Orders"

    # ── Styles ────────────────────────────────────────────────────────────────
    header_fill = PatternFill("solid", fgColor="1565C0")
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    alt_fill = PatternFill("solid", fgColor="E3F2FD")
    cell_font = Font(name="Calibri", size=10)
    center_align = Alignment(horizontal="center", vertical="center")
    left_align   = Alignment(horizontal="left",   vertical="center")
    right_align  = Alignment(horizontal="right",  vertical="center")

    thin = Side(style="thin", color="BDBDBD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── Header row ────────────────────────────────────────────────────────────
    ws.append(list(_HEADERS))
    for col_idx, cell in enumerate(ws[1], start=1):
        cell.fill    = header_fill
        cell.font    = header_font
        cell.alignment = header_align
        cell.border  = border
    ws.row_dimensions[1].height = 22

    # ── Data rows ─────────────────────────────────────────────────────────────
    for row_idx, row in enumerate(rows, start=2):
        fmt = _format_row(row)
        ws.append(list(fmt))

        fill = alt_fill if row_idx % 2 == 0 else None
        for col_idx, cell in enumerate(ws[row_idx], start=1):
            cell.font   = cell_font
            cell.border = border
            if fill:
                cell.fill = fill
            # PO Number: left; numeric cols: right; status/date: center
            if col_idx == 1:
                cell.alignment = left_align
            elif col_idx in (2, 5):          # Vendor ID, Total Amount
                cell.alignment = right_align
            else:
                cell.alignment = center_align
        ws.row_dimensions[row_idx].height = 18

    # ── Column widths ─────────────────────────────────────────────────────────
    col_widths = [20, 12, 14, 14, 16]
    for i, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # ── Freeze header pane ────────────────────────────────────────────────────
    ws.freeze_panes = "A2"

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


# ── Preview endpoint ──────────────────────────────────────────────────────────

@router.get(
    "/purchase-orders/preview",
    summary="Preview report metadata and sample rows",
    response_description="Record count, supported formats, and first 5 rows",
)
def preview_purchase_orders(db: Session = Depends(get_db)):
    """
    Returns report metadata for the UI preview dialog:

    - **total_records** — total count of all purchase orders
    - **format_options** — supported download formats
    - **sample_rows** — first 5 orders (most recent first) as plain objects
    """
    all_rows = _fetch_rows(db)
    total    = len(all_rows)

    sample = []
    for row in all_rows[:5]:
        fmt = _format_row(row)
        sample.append({
            "po_number":    fmt[0],
            "vendor_id":    fmt[1],
            "status":       fmt[2],
            "order_date":   fmt[3],
            "total_amount": fmt[4],
        })

    return {
        "total_records":  total,
        "format_options": ["csv", "pdf", "excel"],
        "sample_rows":    sample,
    }


# ── Export endpoint ────────────────────────────────────────────────────────────

@router.get(
    "/purchase-orders",
    summary="Export purchase orders as CSV, PDF, or Excel",
    response_description="A downloadable CSV, PDF, or Excel file",
    responses={
        200: {"description": "File download"},
    },
)
def export_purchase_orders(
    format: Literal["csv", "pdf", "excel"] = Query(
        default="csv",
        description="Output format: 'csv' (default), 'pdf', or 'excel'",
    ),
    db: Session = Depends(get_db),
):
    """
    Stream all purchase orders as a downloadable file.

    - **format=csv**   — comma-separated values with a header row
    - **format=pdf**   — styled table rendered with ReportLab
    - **format=excel** — styled .xlsx workbook via openpyxl
    """
    rows = _fetch_rows(db)

    if format == "csv":
        csv_buffer = _build_csv(rows)
        return StreamingResponse(
            iter([csv_buffer.getvalue()]),
            media_type="text/csv",
            headers={
                "Content-Disposition": 'attachment; filename="purchase_orders.csv"',
            },
        )

    if format == "excel":
        xlsx_buffer = _build_excel(rows)
        return StreamingResponse(
            xlsx_buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": 'attachment; filename="purchase_orders.xlsx"',
            },
        )

    # format == "pdf"
    pdf_buffer = _build_pdf(rows)
    return StreamingResponse(
        pdf_buffer,
        media_type="application/pdf",
        headers={
            "Content-Disposition": 'attachment; filename="purchase_orders.pdf"',
        },
    )
